import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Palette ──────────────────────────────────────────────────────────────────
BG0      = "111111"   # sheet background (unused – openpyxl can't set sheet bg)
BG1      = "1B1B1B"   # section header rows
BG2      = "242424"   # alternating row
BG3      = "2E2E2E"   # alternating row (lighter)
TXT_H    = "E2E2E2"   # header text
TXT_DIM  = "969696"   # label text
TXT_VAL  = "E2E2E2"   # value text
BLUE     = "4A90D9"   # Claude Code accent
ORANGE   = "E07B39"   # Copilot accent
GREEN    = "5CB85C"   # Cursor accent
AMBER    = "F0AD4E"   # highlight / winner
RED_SOFT = "C0392B"   # warning

def fill(hex_col):
    return PatternFill("solid", fgColor=hex_col)

def font(bold=False, colour=TXT_VAL, size=10, italic=False):
    return Font(bold=bold, color=colour, size=size, italic=italic, name="Calibri")

def border_thin():
    s = Side(style="thin", color="333333")
    return Border(left=s, right=s, top=s, bottom=s)

def border_bottom():
    s = Side(style="thin", color="444444")
    return Border(bottom=s)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def hdr_cell(ws, row, col, value, bg=BG1, fg=TXT_H, bold=True, size=10, align_h="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg)
    c.font = font(bold=bold, colour=fg, size=size)
    c.alignment = align(h=align_h)
    c.border = border_thin()
    return c

def val_cell(ws, row, col, value, bg=BG2, fg=TXT_VAL, bold=False, align_h="right", number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg)
    c.font = font(bold=bold, colour=fg)
    c.alignment = align(h=align_h)
    c.border = border_thin()
    if number_format:
        c.number_format = number_format
    return c

def lbl_cell(ws, row, col, value, bg=BG2, fg=TXT_DIM):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg)
    c.font = font(bold=False, colour=fg)
    c.alignment = align(h="left")
    c.border = border_thin()
    return c

def money(ws, row, col, value, bg=BG2, bold=False, highlight=False):
    fg = AMBER if highlight else TXT_VAL
    c = val_cell(ws, row, col, value, bg=bg, fg=fg, bold=bold, align_h="right", number_format='"$"#,##0')
    return c

def section_row(ws, row, label, ncols, bg=BG1):
    c = ws.cell(row=row, column=1, value=label)
    c.fill = fill(bg)
    c.font = font(bold=True, colour=TXT_H, size=10)
    c.alignment = align(h="left")
    c.border = border_thin()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)

def set_col_width(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def freeze(ws, cell="B2"):
    ws.freeze_panes = cell

def tab_colour(ws, hex_col):
    ws.sheet_properties.tabColor = hex_col


# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 — INPUTS
# ════════════════════════════════════════════════════════════════════════════
ws_in = wb.active
ws_in.title = "Inputs"
tab_colour(ws_in, "4A90D9")
ws_in.sheet_view.showGridLines = False

set_col_width(ws_in, [28, 18, 14, 40])

# Title
r = 1
hdr_cell(ws_in, r, 1, "AI Coding Tool Cost Model — Inputs & Assumptions",
         bg="1B1B1B", fg=TXT_H, bold=True, size=12)
ws_in.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

r = 2
for col in range(1, 5):
    ws_in.cell(row=r, column=col).fill = fill(BG1)

# ── Team ──────────────────────────────────────────────────────────────────
r = 3
section_row(ws_in, r, "TEAM", 4)

data_team = [
    ("Engineers", 10, ""),
    ("Working days / month", 22, ""),
]
for label, val, note in data_team:
    r += 1
    lbl_cell(ws_in, r, 1, label)
    val_cell(ws_in, r, 2, val, align_h="right")
    lbl_cell(ws_in, r, 3, "", bg=BG2)
    lbl_cell(ws_in, r, 4, note, bg=BG2, fg=TXT_DIM)

# ── Task mix ──────────────────────────────────────────────────────────────
r += 1
section_row(ws_in, r, "TASK MIX  (Heavy scenario — 30 tasks/engineer/day)", 4)

hdr_cell(ws_in, r+1, 1, "Tier",       bg=BG1, fg=TXT_DIM, bold=False, size=9)
hdr_cell(ws_in, r+1, 2, "% of Tasks", bg=BG1, fg=TXT_DIM, bold=False, size=9, align_h="right")
hdr_cell(ws_in, r+1, 3, "Input (k tokens)", bg=BG1, fg=TXT_DIM, bold=False, size=9, align_h="right")
hdr_cell(ws_in, r+1, 4, "Output (k tokens)", bg=BG1, fg=TXT_DIM, bold=False, size=9, align_h="right")
r += 1

task_mix = [
    ("Simple  (quick edits, Q&A)",     0.40, 15, 1),
    ("Medium  (bug fixes, features)",  0.35, 35, 3),
    ("Complex (refactors, arch)",       0.15, 60, 8),
    ("Planning (Opus)",                 0.10, 80, 10),
]
bg_cycle = [BG2, BG3]
for i, (tier, pct, inp, out) in enumerate(task_mix):
    r += 1
    bg = bg_cycle[i % 2]
    lbl_cell(ws_in, r, 1, tier, bg=bg)
    val_cell(ws_in, r, 2, pct, bg=bg, align_h="right", number_format="0%")
    val_cell(ws_in, r, 3, inp, bg=bg, align_h="right")
    val_cell(ws_in, r, 4, out, bg=bg, align_h="right")

# ── Model rates ───────────────────────────────────────────────────────────
r += 2
section_row(ws_in, r, "MODEL RATES  ($ per million tokens)", 4)

hdr_cell(ws_in, r+1, 1, "Model",    bg=BG1, fg=TXT_DIM, bold=False, size=9)
hdr_cell(ws_in, r+1, 2, "Input",    bg=BG1, fg=TXT_DIM, bold=False, size=9, align_h="right")
hdr_cell(ws_in, r+1, 3, "Output",   bg=BG1, fg=TXT_DIM, bold=False, size=9, align_h="right")
hdr_cell(ws_in, r+1, 4, "Notes",    bg=BG1, fg=TXT_DIM, bold=False, size=9)
r += 1

rates = [
    ("Claude Haiku 4.5",              1.00,  5.00, "Claude Code — simple tasks"),
    ("Claude Sonnet 4.6",             3.00, 15.00, "Claude Code — medium & complex. Cached input: $0.30/M"),
    ("Claude Opus 4.8 (direct API)", 15.00, 75.00, "Claude Code & Cursor — planning"),
    ("GPT-5 mini (Copilot)",          0.25,  2.00, "Copilot — simple tasks"),
    ("GPT-4.1 (Copilot)",             2.00,  8.00, "Copilot — medium tasks"),
    ("Claude Sonnet via Copilot",     3.00, 15.00, "Copilot — complex tasks"),
    ("Claude Opus via Copilot",       5.00, 25.00, "Copilot — planning. Discounted vs direct API ($15/$75)"),
    ("Cursor Auto",                   1.25,  6.00, "Cursor — simple & medium. Blended/routed model"),
    ("Cursor Max Sonnet",             3.00, 15.00, "Cursor — complex tasks"),
    ("Cursor Max Opus",              15.00, 75.00, "Cursor — planning. Full API rate"),
]
for i, (model, inp, out, note) in enumerate(rates):
    r += 1
    bg = bg_cycle[i % 2]
    lbl_cell(ws_in, r, 1, model, bg=bg)
    val_cell(ws_in, r, 2, inp,  bg=bg, align_h="right", number_format='"$"#,##0.00')
    val_cell(ws_in, r, 3, out,  bg=bg, align_h="right", number_format='"$"#,##0.00')
    lbl_cell(ws_in, r, 4, note, bg=bg, fg=TXT_DIM)

# ── Caching ───────────────────────────────────────────────────────────────
r += 2
section_row(ws_in, r, "CACHING  (Claude Code only — prompt cache hit rate)", 4)
r += 1
lbl_cell(ws_in, r, 1, "Cache hit rate (medium / complex / planning input)")
val_cell(ws_in, r, 2, 0.40, align_h="right", number_format="0%")
lbl_cell(ws_in, r, 3, "")
lbl_cell(ws_in, r, 4, "Cached input billed at $0.30/M vs $3.00/M (Sonnet) or $0.50/M vs $15/M (Opus)", fg=TXT_DIM)

# ── Plan subscriptions ────────────────────────────────────────────────────
r += 2
section_row(ws_in, r, "PLAN SUBSCRIPTIONS  (per month, 10 engineers)", 4)

hdr_cell(ws_in, r+1, 1, "Tool",             bg=BG1, fg=TXT_DIM, bold=False, size=9)
hdr_cell(ws_in, r+1, 2, "Monthly Base",     bg=BG1, fg=TXT_DIM, bold=False, size=9, align_h="right")
hdr_cell(ws_in, r+1, 3, "Included Credits", bg=BG1, fg=TXT_DIM, bold=False, size=9, align_h="right")
hdr_cell(ws_in, r+1, 4, "Notes",            bg=BG1, fg=TXT_DIM, bold=False, size=9)
r += 1

plans = [
    ("Copilot Business",  190,  190, "$19/user — credits = subscription cost"),
    ("Cursor Teams",      400,  200, "$40/user — $20 platform fee + $20 credits per user"),
    ("Claude Code",         0,    0, "API billing only in this model (no flat seat fee assumed)"),
]
for i, (tool, base, credits, note) in enumerate(plans):
    r += 1
    bg = bg_cycle[i % 2]
    lbl_cell(ws_in, r, 1, tool, bg=bg)
    money(ws_in, r, 2, base, bg=bg)
    money(ws_in, r, 3, credits, bg=bg)
    lbl_cell(ws_in, r, 4, note, bg=bg, fg=TXT_DIM)

freeze(ws_in, "A2")


# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 — DETAILED MODEL (Heavy + Opus, primary scenario)
# ════════════════════════════════════════════════════════════════════════════
ws_d = wb.create_sheet("Heavy + Opus (Detail)")
tab_colour(ws_d, "E07B39")
ws_d.sheet_view.showGridLines = False

set_col_width(ws_d, [30, 16, 16, 16, 16, 16, 16, 16])

# Engineers=10, days=22, tasks/day=30
ENG = 10; DAYS = 22; TPD = 30
TOTAL_TASKS = ENG * DAYS * TPD   # 6600

mix    = [0.40, 0.35, 0.15, 0.10]
inp_k  = [15,   35,   60,   80  ]
out_k  = [1,    3,    8,    10  ]
tiers  = ["Simple", "Medium", "Complex", "Planning (Opus)"]

tasks_per_tier = [round(TOTAL_TASKS * m) for m in mix]

r = 1
hdr_cell(ws_d, r, 1, "Heavy Scenario — Mixed Models + Opus Planning  |  10 engineers · 30 tasks/day · 22 days",
         bg=BG1, fg=TXT_H, bold=True, size=11)
ws_d.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

r = 2
for col in range(1, 9):
    ws_d.cell(row=r, column=col).fill = fill(BG1)

# ── Column headers ──────────────────────────────────────────────────────
r = 3
cols_d = ["Tier", "Tasks", "Input (M tok)", "Output (M tok)",
          "Claude Code ($)", "Copilot ($)", "Cursor ($)", "Notes"]
bgs_d  = [BG1, BG1, BG1, BG1, BLUE[:6] and "1A3A5C", "2A2010", "0D2E1A", BG1]
tool_bgs = ["1A3A5C", "2A2010", "0D2E1A"]

for ci, label in enumerate(cols_d, 1):
    bg = BG1
    c = ws_d.cell(row=r, column=ci, value=label)
    c.fill = fill(bg)
    c.font = font(bold=True, colour=TXT_DIM, size=9)
    c.alignment = align(h="right" if ci > 1 else "left")
    c.border = border_thin()

# ── Per-tier rows ────────────────────────────────────────────────────────
# Model rates
CC_rates  = [(1.00, 5.00), (3.00, 15.00), (3.00, 15.00), (15.00, 75.00)]
COP_rates = [(0.25, 2.00), (2.00,  8.00), (3.00, 15.00), ( 5.00, 25.00)]
CUR_rates = [(1.25, 6.00), (1.25,  6.00), (3.00, 15.00), (15.00, 75.00)]
CC_models  = ["Haiku",    "Sonnet",  "Sonnet",      "Opus (direct $15/$75)"]
COP_models = ["GPT-5mini","GPT-4.1", "Sonnet",      "Opus (Copilot $5/$25)"]
CUR_models = ["Auto",     "Auto",    "Max Sonnet",  "Max Opus (direct $15/$75)"]
CACHE_RATE = 0.40
SONNET_CACHE_IN = 0.30
OPUS_CACHE_IN   = 0.50

cc_tier_costs = []
cop_tier_costs = []
cur_tier_costs = []

for i, tier in enumerate(tiers):
    r += 1
    bg = bg_cycle[i % 2]
    tasks = tasks_per_tier[i]
    inp_m = tasks * inp_k[i] / 1_000
    out_m = tasks * out_k[i] / 1_000

    # Claude Code
    cc_in_r, cc_out_r = CC_rates[i]
    if i == 0:  # simple — no caching
        cc_cost = inp_m * cc_in_r + out_m * cc_out_r
    elif i in (1, 2):  # medium/complex — Sonnet with caching
        cc_cached_in = inp_m * CACHE_RATE * SONNET_CACHE_IN
        cc_fresh_in  = inp_m * (1 - CACHE_RATE) * cc_in_r
        cc_cost = cc_cached_in + cc_fresh_in + out_m * cc_out_r
    else:  # planning — Opus with caching
        cc_cached_in = inp_m * CACHE_RATE * OPUS_CACHE_IN
        cc_fresh_in  = inp_m * (1 - CACHE_RATE) * cc_in_r
        cc_cost = cc_cached_in + cc_fresh_in + out_m * cc_out_r

    # Copilot
    cop_in_r, cop_out_r = COP_rates[i]
    cop_cost = inp_m * cop_in_r + out_m * cop_out_r

    # Cursor
    cur_in_r, cur_out_r = CUR_rates[i]
    cur_cost = inp_m * cur_in_r + out_m * cur_out_r

    cc_tier_costs.append(cc_cost)
    cop_tier_costs.append(cop_cost)
    cur_tier_costs.append(cur_cost)

    note = f"CC:{CC_models[i]}  COP:{COP_models[i]}  CUR:{CUR_models[i]}"

    lbl_cell(ws_d, r, 1, tier, bg=bg)
    val_cell(ws_d, r, 2, tasks, bg=bg, align_h="right")
    val_cell(ws_d, r, 3, round(inp_m, 1), bg=bg, align_h="right", number_format="#,##0.0")
    val_cell(ws_d, r, 4, round(out_m, 1), bg=bg, align_h="right", number_format="#,##0.0")
    money(ws_d, r, 5, round(cc_cost),  bg=bg)
    money(ws_d, r, 6, round(cop_cost), bg=bg)
    money(ws_d, r, 7, round(cur_cost), bg=bg)
    lbl_cell(ws_d, r, 8, note, bg=bg, fg=TXT_DIM)

# ── Totals ───────────────────────────────────────────────────────────────
r += 1
section_row(ws_d, r, "", 8, bg=BG1)

r += 1
hdr_cell(ws_d, r, 1, "Total token cost", bg=BG1)
ws_d.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
for col in range(1, 5):
    ws_d.cell(row=r, column=col).fill = fill(BG1)
    ws_d.cell(row=r, column=col).font = font(bold=True, colour=TXT_H)
    ws_d.cell(row=r, column=col).border = border_thin()

cc_token_total  = sum(cc_tier_costs)
cop_token_total = sum(cop_tier_costs)
cur_token_total = sum(cur_tier_costs)

money(ws_d, r, 5, round(cc_token_total),  bg=BG1, bold=True)
money(ws_d, r, 6, round(cop_token_total), bg=BG1, bold=True)
money(ws_d, r, 7, round(cur_token_total), bg=BG1, bold=True)
lbl_cell(ws_d, r, 8, "", bg=BG1)

# ── Platform / subscription adjustments ─────────────────────────────────
r += 1
hdr_cell(ws_d, r, 1, "Plan base cost", bg=BG2)
ws_d.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
for col in range(1, 5):
    ws_d.cell(row=r, column=col).fill = fill(BG2)
    ws_d.cell(row=r, column=col).font = font(colour=TXT_DIM)
    ws_d.cell(row=r, column=col).border = border_thin()
money(ws_d, r, 5, 0,   bg=BG2)
money(ws_d, r, 6, 190, bg=BG2)
money(ws_d, r, 7, 400, bg=BG2)
lbl_cell(ws_d, r, 8, "Copilot: $19×10. Cursor: $40×10 ($200 platform + $200 credits)", bg=BG2, fg=TXT_DIM)

r += 1
hdr_cell(ws_d, r, 1, "Included credits", bg=BG2)
ws_d.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
for col in range(1, 5):
    ws_d.cell(row=r, column=col).fill = fill(BG2)
    ws_d.cell(row=r, column=col).font = font(colour=TXT_DIM)
    ws_d.cell(row=r, column=col).border = border_thin()
money(ws_d, r, 5, 0,   bg=BG2)
money(ws_d, r, 6, 190, bg=BG2)
money(ws_d, r, 7, 200, bg=BG2)
lbl_cell(ws_d, r, 8, "Credits offset token costs to this amount", bg=BG2, fg=TXT_DIM)

r += 1
hdr_cell(ws_d, r, 1, "Overage (token cost − credits)", bg=BG2)
ws_d.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
for col in range(1, 5):
    ws_d.cell(row=r, column=col).fill = fill(BG2)
    ws_d.cell(row=r, column=col).font = font(colour=TXT_DIM)
    ws_d.cell(row=r, column=col).border = border_thin()
cc_overage  = max(0, cc_token_total  - 0)
cop_overage = max(0, cop_token_total - 190)
cur_overage = max(0, cur_token_total - 200)
money(ws_d, r, 5, round(cc_overage),  bg=BG2)
money(ws_d, r, 6, round(cop_overage), bg=BG2)
money(ws_d, r, 7, round(cur_overage), bg=BG2)
lbl_cell(ws_d, r, 8, "", bg=BG2)

r += 1
cc_total  = round(cc_token_total  + 0)
cop_total = round(cop_token_total)   # base already = credits so total = token cost
cur_total = round(cur_token_total - 200 + 400)
winner = min(cc_total, cop_total, cur_total)

hdr_cell(ws_d, r, 1, "TOTAL MONTHLY BILL", bg=BG1, bold=True)
ws_d.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
for col in range(1, 5):
    ws_d.cell(row=r, column=col).fill = fill(BG1)
    ws_d.cell(row=r, column=col).font = font(bold=True, colour=TXT_H)
    ws_d.cell(row=r, column=col).border = border_thin()
money(ws_d, r, 5, cc_total,  bg=BG1, bold=True, highlight=(cc_total==winner))
money(ws_d, r, 6, cop_total, bg=BG1, bold=True, highlight=(cop_total==winner))
money(ws_d, r, 7, cur_total, bg=BG1, bold=True, highlight=(cur_total==winner))
lbl_cell(ws_d, r, 8, "★ = lowest cost", bg=BG1, fg=AMBER)

freeze(ws_d, "B4")


# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 — SUMMARY ACROSS ALL SCENARIOS
# ════════════════════════════════════════════════════════════════════════════
ws_s = wb.create_sheet("Summary")
tab_colour(ws_s, "5CB85C")
ws_s.sheet_view.showGridLines = False
set_col_width(ws_s, [36, 18, 18, 18, 32])

r = 1
hdr_cell(ws_s, r, 1, "AI Coding Tool Cost Model — Summary  |  10 Engineers",
         bg=BG1, fg=TXT_H, bold=True, size=12)
ws_s.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

r = 2
for col in range(1, 6):
    ws_s.cell(row=r, column=col).fill = fill(BG1)

r = 3
for ci, h in enumerate(["Scenario", "Claude Code", "Copilot", "Cursor", "Key observation"], 1):
    c = ws_s.cell(row=r, column=ci, value=h)
    c.fill = fill(BG1)
    c.font = font(bold=True, colour=TXT_DIM, size=9)
    c.alignment = align(h="right" if 1 < ci < 5 else "left")
    c.border = border_thin()

scenarios = [
    # label,                           CC,    COP,   CUR,   note
    ("Light  (10 tasks/day, Sonnet/GPT-4.1/Auto)",
                                        231,   190,   400,
     "Copilot cheapest. Cursor's $400 floor hurts at low usage."),
    ("Medium  (20 tasks/day, Sonnet/GPT-4.1/Auto)",
                                        594,   370,   444,
     "Copilot still cheapest. Cursor Auto catches up."),
    ("Heavy  (30 tasks/day, single model — Sonnet/GPT-4.1/Auto)",
                                       1287,   792,   728,
     "Cursor Auto cheapest. Copilot gains from GPT-4.1 rate."),
    ("Heavy, mixed models, no Opus  (40/40/20 split)",
                                        600,   659,   824,
     "Claude Code cheapest with caching. Cursor floor expensive."),
    ("Heavy, mixed models + Opus planning  (40/35/15/10 split)",
                                       1500,   958,  1992,
     "Copilot dominant — Opus discount ($5/M vs $15/M direct)."),
]

for i, (label, cc, cop, cur, note) in enumerate(scenarios):
    r += 1
    bg = bg_cycle[i % 2]
    winner = min(cc, cop, cur)
    lbl_cell(ws_s, r, 1, label, bg=bg)
    money(ws_s, r, 2, cc,  bg=bg, highlight=(cc==winner))
    money(ws_s, r, 3, cop, bg=bg, highlight=(cop==winner))
    money(ws_s, r, 4, cur, bg=bg, highlight=(cur==winner))
    lbl_cell(ws_s, r, 5, note, bg=bg, fg=TXT_DIM)

r += 2
section_row(ws_s, r, "KEY FINDINGS", 5)
findings = [
    "Copilot's discounted Opus rate ($5/M vs $15/M direct) is the single biggest swing factor. Teams using Opus for planning should default to Copilot.",
    "Claude Code's prompt caching (est. 40% hit rate) significantly reduces real costs — raw token calculations overstate actual bills by ~30–40%.",
    "Cursor's $400 floor makes it expensive at light usage. Its Auto mode is cost-efficient at medium-heavy usage without Opus.",
    "No single tool wins across all scenarios. Usage pattern — especially Opus frequency — determines optimal choice.",
    "Copilot's Opus discount is commercially unusual (1/3 of retail). Treat it as a potential short-term advantage, not a permanent structural feature.",
]
for i, f in enumerate(findings):
    r += 1
    bg = bg_cycle[i % 2]
    c = ws_s.cell(row=r, column=1, value=f"  {i+1}.  {f}")
    c.fill = fill(bg)
    c.font = font(colour=TXT_VAL, italic=True)
    c.alignment = align(h="left", wrap=True)
    c.border = border_thin()
    ws_s.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws_s.row_dimensions[r].height = 30

freeze(ws_s, "A4")


# ════════════════════════════════════════════════════════════════════════════
# SHEET 4 — SENSITIVITY (Opus % of tasks)
# ════════════════════════════════════════════════════════════════════════════
ws_sens = wb.create_sheet("Sensitivity — Opus %")
tab_colour(ws_sens, "F0AD4E")
ws_sens.sheet_view.showGridLines = False
set_col_width(ws_sens, [24, 18, 18, 18, 32])

r = 1
hdr_cell(ws_sens, r, 1, "Sensitivity: Opus Planning % vs Monthly Cost  |  Heavy scenario (30 tasks/day, 10 engineers)",
         bg=BG1, fg=TXT_H, bold=True, size=11)
ws_sens.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

r = 2
for col in range(1, 6):
    ws_sens.cell(row=r, column=col).fill = fill(BG1)

r = 3
for ci, h in enumerate(["Opus % of tasks", "Claude Code", "Copilot", "Cursor", "Cheapest tool"], 1):
    c = ws_sens.cell(row=r, column=ci, value=h)
    c.fill = fill(BG1)
    c.font = font(bold=True, colour=TXT_DIM, size=9)
    c.alignment = align(h="right" if 1 < ci < 5 else "left")
    c.border = border_thin()

def calc_costs(opus_pct):
    # Remaining split across simple/medium/complex proportionally from 40/35/15
    remaining = 1.0 - opus_pct
    base_total = 0.40 + 0.35 + 0.15
    sp = (0.40 / base_total) * remaining
    mp = (0.35 / base_total) * remaining
    cp = (0.15 / base_total) * remaining

    mix_s = [sp, mp, cp, opus_pct]
    tasks_s = [round(TOTAL_TASKS * p) for p in mix_s]

    cc_cost = 0; cop_cost = 0; cur_cost = 0
    for i in range(4):
        inp_m = tasks_s[i] * inp_k[i] / 1_000
        out_m = tasks_s[i] * out_k[i] / 1_000
        # CC
        cc_in_r, cc_out_r = CC_rates[i]
        if i == 0:
            cc_cost += inp_m * cc_in_r + out_m * cc_out_r
        elif i in (1, 2):
            cc_cost += inp_m * CACHE_RATE * SONNET_CACHE_IN + inp_m * (1-CACHE_RATE) * cc_in_r + out_m * cc_out_r
        else:
            cc_cost += inp_m * CACHE_RATE * OPUS_CACHE_IN + inp_m * (1-CACHE_RATE) * cc_in_r + out_m * cc_out_r
        # Copilot
        cop_in_r, cop_out_r = COP_rates[i]
        cop_cost += inp_m * cop_in_r + out_m * cop_out_r
        # Cursor
        cur_in_r, cur_out_r = CUR_rates[i]
        cur_cost += inp_m * cur_in_r + out_m * cur_out_r

    cc_total  = round(cc_cost)
    cop_total = round(cop_cost)
    cur_total = round(cur_cost - 200 + 400)
    return cc_total, cop_total, cur_total

tool_names = {0: "Claude Code", 1: "Copilot", 2: "Cursor"}

for i, opus_pct in enumerate([0.00, 0.05, 0.10, 0.15, 0.20, 0.25, 0.30]):
    r += 1
    bg = bg_cycle[i % 2]
    cc, cop, cur = calc_costs(opus_pct)
    winner_idx = [cc, cop, cur].index(min(cc, cop, cur))
    winner_name = tool_names[winner_idx]

    val_cell(ws_sens, r, 1, opus_pct, bg=bg, align_h="right", number_format="0%")
    money(ws_sens, r, 2, cc,  bg=bg, highlight=(winner_idx==0))
    money(ws_sens, r, 3, cop, bg=bg, highlight=(winner_idx==1))
    money(ws_sens, r, 4, cur, bg=bg, highlight=(winner_idx==2))
    lbl_cell(ws_sens, r, 5, winner_name, bg=bg, fg=AMBER if True else TXT_DIM)

r += 2
c = ws_sens.cell(row=r, column=1,
    value="Highlighted cells = lowest cost for that Opus %. "
          "Copilot's Opus discount makes it cheapest once Opus tasks exceed ~5% of daily workload.")
c.fill = fill(BG1)
c.font = font(colour=TXT_DIM, italic=True, size=9)
c.alignment = align(h="left", wrap=True)
c.border = border_thin()
ws_sens.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
ws_sens.row_dimensions[r].height = 28

freeze(ws_sens, "A4")

# ── Save ─────────────────────────────────────────────────────────────────────
out_path = r"C:\Users\EdYau\reinvently.www\ai-coding-tool-cost-model.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
